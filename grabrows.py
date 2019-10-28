#==========================================================================================================================
# Authors: Dr. Eaglin, Andrew Polemeni
# Organization: Daytona State College
# Date Written: October 24, 2019
# Program Purpose: To automate course check downs for student progress.
#==========================================================================================================================

# Import statements
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
import main

#==========================================================================================================================

#Grab the first excel file which will be the Query Report
# Removes all columns and values that should not be used
#Get the Query

def GetStudentData(query_filename):
    # Answers a DataFrame
    # query = file name of excel spreadsheet containing query info
    # sid is student id
    df1 = pd.read_excel(query_filename, sheet_name='sheet1', usecols="A, B, C, D, E, F, G, M") #df = dataframe variable and import file here
    df1.dropna(inplace=True) #drop cells with NaN
    df1.drop(df1.loc[df1['Grade']=='F'].index, inplace=True) # Delete grades that are equal to F
    df1.drop(df1.loc[df1['Grade']=='FN'].index, inplace=True) # Delete grades that are equal to FN
    df1.drop(df1.loc[df1['Grade']=='W'].index, inplace=True) # Delete grades that are equal to W
    df1.drop(df1.loc[df1['Grade']=='D'].index, inplace=True) # Delete grades that are equal to D
    df1.drop(df1.loc[df1['Grade']=='I'].index, inplace=True) # Delete grades that are equal to I
    return df1

 

def FilterForStudent(df1, sid):
    #Drop rows not needed or used
    # Answers a DataFrame
    # sid is student id
    # Creates a new copy of
    df2 = df1.copy(deep = True)
    df2.drop(df2.loc[df1["ID"] != int(sid)].index, inplace=True)

    return df2   

#==========================================================================================================================
# File to be pasted into - this is a blank workbook
def CreateStudentFile(df):

    file = "BSET"
    if (df['Acad Plan'].iloc[0] == 633400):

        file = "BSIT"

    if (df['Acad Plan'].iloc[0] == 633300): # Find the dataframe by string and equal to program number

        file = "BSIT"

    wb2 = openpyxl.load_workbook(file + ".xlsx")
    ws2 = wb2.active

#==========================================================================================================================
# We now simply go through each column of the BSET Checkdown
# and compare against courses taken
    for i in range(1,60):
        course = str(ws2.cell(row=i, column=3).value)[0:7]
        for row in df.iterrows():
            taken = row[1]['Subject'] + str(row[1]['Catalog'])
            name = row[1]['First Name'] + "_" + str(row[1]['Last'])

            if (course == taken):

                ws2.cell(row=i, column=4).value = str(row[1]['Term'])
                ws2.cell(row=i, column=4).fill = PatternFill(fgColor=YELLOW, fill_type = "solid")


#Once the sheet is filled out

    file_to_save = file + "_" + name + '.xlsx'
    wb2.save(file_to_save)
    return file_to_save

# Use of the functions
# Get the ID of student for report

df1 = GetStudentData("Query.xlsx") #CREATE A VARIABLE FOR THE QUERY SHEET
ids = df1['ID'].unique().tolist()


# This will generate a spreadsheet for everyone
def GenerateAllSpreadsheets():
    for sid in ids:
        df2 = FilterForStudent(df1, sid)  
        file = CreateStudentFile(df2)
        print("Student Information save to " + file)

    file = CreateStudentFile(df2)
    print(df2['Acad Plan'].iloc[0])
    file = CreateStudentFile(df2)
